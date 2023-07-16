import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx') # wb=workbook, which is an object
sheet = wb['Sayfa1']#it is Sheet1 if computer's language is in English
cell = sheet['a1']
cell = sheet.cell(1,1) #cell(row,column), which can be used instead of sheet['a1']
# print(cell.value)
# print(sheet.max_row) #returns 4, meaning there are 4 rows

"""for row in range(1, sheet.max_row + 1):
    print(row)"""
"""for row in range(2, sheet.max_row + 1): #started with 2 for showing only values, exempting column name
    cell = sheet.cell(row, 3)
    print(cell.value)"""
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4
          )
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'f2')

wb.save('transactions2.xlsx')